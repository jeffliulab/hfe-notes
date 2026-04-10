from __future__ import annotations

import html
import io
import json
import posixpath
import re
import shutil
from urllib.parse import quote
import zipfile
from collections import Counter, OrderedDict, defaultdict
from pathlib import Path, PurePosixPath
import xml.etree.ElementTree as ET

import fitz
from openpyxl import load_workbook
from page_content import PAGE_CONTENT


REPO_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = REPO_ROOT.parent
DOCS_DIR = REPO_ROOT / "docs"
DATA_DIR = REPO_ROOT / "data"
ASSET_ROOT = DOCS_DIR / "assets" / "source_files"
VISUAL_ROOT = DOCS_DIR / "assets" / "visuals"
SITE_URL = "https://jeffliulab.github.io/hfe-notes/"

PPT_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
REL_NS_URI = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
SUPPORTED_VISUAL_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp"}

SOURCE_ARCHIVES = [
    {
        "zip_name": "ENP 111 Use-related Risks.zip",
        "asset_dir": "ENP_111_Use_related_Risks",
    },
    {
        "zip_name": "Lectures_Spring____6_export.zip",
        "asset_dir": "Lectures_Spring_2026",
    },
]

SECTIONS = OrderedDict(
    {
        "HFE_Foundations": {
            "zh_nav": "HFE基础",
            "en_nav": "HFE Foundations",
            "zh_index_nav": "HFE基础总览",
            "en_index_nav": "HFE Foundations Overview",
            "zh_title": "HFE基础",
            "en_title": "HFE Foundations",
            "zh_intro": "本分区聚焦课程导论、人因定义、人为失误框架与 Swiss Cheese 模型，帮助读者建立整个课程的语言体系与分析视角。",
            "en_intro": "This section establishes the language of the course through the course overview, core human-factors definitions, human-error frameworks, and the Swiss Cheese model.",
            "zh_card": "课程导论、人因定义、人为失误框架与 Swiss Cheese 模型。",
            "en_card": "Course framing, HF definitions, human-error frameworks, and the Swiss Cheese model.",
        },
        "HFE_Risk_Methods": {
            "zh_nav": "风险方法",
            "en_nav": "Risk Methods",
            "zh_index_nav": "风险方法总览",
            "en_index_nav": "Risk Methods Overview",
            "zh_title": "风险方法",
            "en_title": "Risk Methods",
            "zh_intro": "本分区整理课程中的分析方法论，包括调查流程、错误分析、任务分析、URRA、Known Problem Analysis 与事件树思路。",
            "en_intro": "This section organizes the analytic methods in the course, including investigation flow, error analysis, task analysis, URRA, known problem analysis, and event-tree thinking.",
            "zh_card": "调查流程、错误分析、任务分析、URRA 与事件树方法。",
            "en_card": "Investigation flow, error analysis, task analysis, URRA, and event-tree methods.",
        },
        "HFE_Medical_Devices": {
            "zh_nav": "医疗器械",
            "en_nav": "Medical Devices",
            "zh_index_nav": "医疗器械总览",
            "en_index_nav": "Medical Devices Overview",
            "zh_title": "医疗器械",
            "en_title": "Medical Devices",
            "zh_intro": "本分区将 ISO 14971、医疗器械 use error、URRA 实践与 EpiPen 工作簿整合成一套面向医疗器械风险管理的知识链路。",
            "en_intro": "This section connects ISO 14971, medical-device use error, URRA practice, and the EpiPen workbook into a coherent medical-device risk-management track.",
            "zh_card": "ISO 14971、use error、医疗器械 URRA 与 EpiPen 工作簿。",
            "en_card": "ISO 14971, use error, medical-device URRA, and the EpiPen workbook.",
        },
        "HFE_Aviation_Automation": {
            "zh_nav": "航空与自动化",
            "en_nav": "Aviation & Automation",
            "zh_index_nav": "航空与自动化总览",
            "en_index_nav": "Aviation & Automation Overview",
            "zh_title": "航空与自动化",
            "en_title": "Aviation and Automation",
            "zh_intro": "本分区覆盖航空环境中的自动化、人机界面、CRM、显示与告警、检查单及自动化车辆等典型应用。",
            "en_intro": "This section covers aviation-flavored automation, HMI, CRM, displays and alerts, checklists, procedures, and automated vehicles.",
            "zh_card": "航空自动化、CRM、显示与告警、检查单与自动化车辆。",
            "en_card": "Aviation automation, CRM, displays and alerts, checklists, and automated vehicles.",
        },
        "HFE_Human_Performance": {
            "zh_nav": "人的表现",
            "en_nav": "Human Performance",
            "zh_index_nav": "人的表现总览",
            "en_index_nav": "Human Performance Overview",
            "zh_title": "人的表现",
            "en_title": "Human Performance",
            "zh_intro": "本分区围绕注意、疲劳、压力、决策、情境意识与空间定向错觉，整理影响操作者表现的关键心理与生理变量。",
            "en_intro": "This section focuses on attention, fatigue, stress, decision making, situation awareness, and spatial disorientation as key drivers of human performance.",
            "zh_card": "注意、疲劳、压力、决策、情境意识与空间定向。",
            "en_card": "Attention, fatigue, stress, decision making, situation awareness, and spatial orientation.",
        },
        "HFE_Cases_Ethics": {
            "zh_nav": "案例与伦理",
            "en_nav": "Cases & Ethics",
            "zh_index_nav": "案例与伦理总览",
            "en_index_nav": "Cases & Ethics Overview",
            "zh_title": "案例与伦理",
            "en_title": "Cases and Ethics",
            "zh_intro": "本分区收束到案例分析与伦理讨论，涵盖 operational risk、Cardosi 案例、F16 prompts 与 Boeing 737 Max。",
            "en_intro": "This section closes with case studies and ethics, including operational risk, the Cardosi case, F16 prompts, and Boeing 737 Max.",
            "zh_card": "Operational risk、Cardosi、F16 分析 prompts 与 Boeing 737 Max。",
            "en_card": "Operational risk, Cardosi, F16 prompts, and Boeing 737 Max.",
        },
    }
)

PAGES = [
    {
        "slug": "course_overview",
        "section": "HFE_Foundations",
        "zh_nav": "课程导论",
        "en_nav": "Course Overview",
        "zh_title": "课程导论与学习地图",
        "en_title": "Course Overview and Learning Map",
        "zh_pitch": "本页用课程导论材料建立整门课的范围、节奏与应用背景，帮助读者先看到人因工程在航空与复杂系统中的问题空间。",
        "en_pitch": "This page uses the course introduction to frame the scope, pacing, and application domain of the class before the later methods and case studies appear.",
        "angles_zh": ["课程要解决什么系统层问题", "为什么航空案例适合做人因分析入口", "后续主题如何从导论向方法与案例展开"],
        "angles_en": ["What system-level problems the course targets", "Why aviation is a strong entry point for HFE", "How later topics branch from the introductory frame"],
        "source_files": ["Course intro 1-14-26.pdf"],
        "cross_refs": ["human_factors_intro", "error_analysis_methods"],
    },
    {
        "slug": "human_factors_intro",
        "section": "HFE_Foundations",
        "zh_nav": "人因导论",
        "en_nav": "Introduction to Human Factors",
        "zh_title": "人因工程导论",
        "en_title": "Introduction to Human Factors",
        "zh_pitch": "本页围绕 human factors 的定义、设计目标与系统视角，整理人因工程如何把人的能力与限制带回系统设计。",
        "en_pitch": "This page organizes the core definition, design objective, and systems view of human factors as an engineering discipline.",
        "angles_zh": ["Human factors 的学科定义", "优化福祉与系统性能的双目标", "把人视为系统组成部分而不是误差来源"],
        "angles_en": ["The disciplinary definition of human factors", "The dual goals of well-being and performance", "Treating humans as part of the system rather than just error sources"],
        "source_files": ["01 Intro to HF and Risk (1).pptx"],
        "cross_refs": ["human_error_frameworks", "swiss_cheese"],
    },
    {
        "slug": "human_error_frameworks",
        "section": "HFE_Foundations",
        "zh_nav": "人为失误框架",
        "en_nav": "Human Error Frameworks",
        "zh_title": "人为失误框架",
        "en_title": "Human Error Frameworks",
        "zh_pitch": "本页整合课程里对 human error 的引入与框架化讲解，把 slips、mistakes、violations 与多层次失误模型放到同一个分析坐标系里。",
        "en_pitch": "This page integrates the course treatment of human error and places slips, mistakes, violations, and multi-level models in a single analytic frame.",
        "angles_zh": ["错误类型如何区分", "个体层与系统层解释如何互补", "框架为什么会影响后续风险分析方法"],
        "angles_en": ["How error categories are distinguished", "How individual and system explanations complement each other", "Why frameworks matter for later risk methods"],
        "source_files": ["02 Intro to Human Error (1).pptx", "03 HE Frameworks (2).pptx"],
        "cross_refs": ["human_factors_intro", "swiss_cheese", "error_analysis_methods"],
    },
    {
        "slug": "swiss_cheese",
        "section": "HFE_Foundations",
        "zh_nav": "Swiss Cheese 模型",
        "en_nav": "Swiss Cheese Model",
        "zh_title": "Swiss Cheese 模型",
        "en_title": "The Swiss Cheese Model",
        "zh_pitch": "本页用 Swiss Cheese 模型把防御层、薄弱环节与事故路径串起来，作为从单点失误转向系统防护设计的重要桥梁。",
        "en_pitch": "This page uses the Swiss Cheese model to connect barriers, vulnerabilities, and accident trajectories as a bridge from single-error thinking to system defense design.",
        "angles_zh": ["事故为什么穿透多层防线", "防御层与漏洞如何同时存在", "系统改进为什么比责怪操作者更关键"],
        "angles_en": ["Why accidents penetrate multiple defenses", "How barriers and holes coexist", "Why system redesign matters more than blaming operators"],
        "source_files": ["05 Swiss Cheese Model.pdf"],
        "cross_refs": ["human_error_frameworks", "operational_risk"],
    },
    {
        "slug": "error_analysis_methods",
        "section": "HFE_Risk_Methods",
        "zh_nav": "错误分析方法",
        "en_nav": "Error Analysis Methods",
        "zh_title": "错误分析与调查流程",
        "en_title": "Error Analysis and Investigation Flow",
        "zh_pitch": "本页把调查流程、NTSB 取证逻辑与错误分析方法放在一起，说明事故与事件分析如何从事实收集走向可执行的因果解释。",
        "en_pitch": "This page joins the investigative process, NTSB logic, and error-analysis methods to show how event analysis moves from evidence gathering to causal explanation.",
        "angles_zh": ["调查流程的阶段划分", "NTSB 风格的事实到原因链条", "错误分析如何支撑后续风险控制"],
        "angles_en": ["The phases of an investigation", "The NTSB-style path from facts to causes", "How error analysis supports later risk controls"],
        "source_files": ["Sp26_Investigative Process-NTSB_20260121.pdf", "Sp26_ErrorAnalysisMethods_20260121.pdf"],
        "cross_refs": ["human_error_frameworks", "task_analysis", "operational_risk"],
    },
    {
        "slug": "task_analysis",
        "section": "HFE_Risk_Methods",
        "zh_nav": "任务分析",
        "en_nav": "Task Analysis",
        "zh_title": "任务分析",
        "en_title": "Task Analysis",
        "zh_pitch": "本页围绕 task analysis 的拆解逻辑，整理如何把使用流程变成可观察、可评估、可映射风险的任务与子任务结构。",
        "en_pitch": "This page explains how task analysis decomposes use flow into observable tasks and subtasks that can later be evaluated for risk.",
        "angles_zh": ["任务拆解为什么是风险分析前提", "subtask 与 use step 如何定义", "如何从任务流转向 use error 与 mitigation"],
        "angles_en": ["Why task decomposition is foundational for risk analysis", "How subtasks and use steps are defined", "How task flow leads into use error and mitigation"],
        "source_files": ["08 Task Analysis.pptx"],
        "cross_refs": ["urra_methods", "epipen_workbook", "medical_device_use_errors"],
    },
    {
        "slug": "urra_methods",
        "section": "HFE_Risk_Methods",
        "zh_nav": "URRA方法",
        "en_nav": "URRA Methods",
        "zh_title": "URRA 方法",
        "en_title": "URRA Methods",
        "zh_pitch": "本页整理 Use-Related Risk Analysis 的写法、输入输出与判定思路，突出 URRA 在人因风险管理中的中枢作用。",
        "en_pitch": "This page captures how Use-Related Risk Analysis is structured, what inputs and outputs it requires, and why it sits at the center of HFE risk work.",
        "angles_zh": ["URRA 的目标与文档结构", "风险链条如何从任务与错误导出", "URRA 如何为验证与控制措施服务"],
        "angles_en": ["The purpose and document structure of URRA", "How the risk chain is derived from tasks and errors", "How URRA supports validation and controls"],
        "source_files": ["10 Write URRA.pptx"],
        "cross_refs": ["task_analysis", "medical_device_urra", "epipen_workbook"],
    },
    {
        "slug": "known_problem_and_event_tree",
        "section": "HFE_Risk_Methods",
        "zh_nav": "Known Problem 与事件树",
        "en_nav": "Known Problems and Event Trees",
        "zh_title": "Known Problem Analysis 与事件树",
        "en_title": "Known Problem Analysis and Event Trees",
        "zh_pitch": "本页把 known problem analysis 与 event tree 思维放在一起，展示如何从已知问题、先前经验和事件演化来补强风险分析。",
        "en_pitch": "This page pairs known problem analysis with event-tree thinking to show how prior issues and event evolution enrich risk analysis.",
        "angles_zh": ["已知问题为何值得单独建模", "事件树如何展开事件后果", "补充分析如何帮助发现遗漏风险"],
        "angles_en": ["Why known problems deserve explicit modeling", "How event trees expand downstream consequences", "How supplemental analysis catches missed risks"],
        "source_files": ["11 Known Problem Analysis.pptx", "Event Tree Supplemental.pptx"],
        "cross_refs": ["error_analysis_methods", "urra_methods", "operational_risk"],
    },
    {
        "slug": "iso_14971",
        "section": "HFE_Medical_Devices",
        "zh_nav": "ISO 14971",
        "en_nav": "ISO 14971",
        "zh_title": "ISO 14971 与医疗器械风险管理",
        "en_title": "ISO 14971 and Medical Device Risk Management",
        "zh_pitch": "本页围绕 ISO 14971 标准文本与教学材料，整理医疗器械风险管理的基本流程、术语与与人因工作的接口。",
        "en_pitch": "This page uses the course material on ISO 14971 to summarize the medical-device risk-management process, terminology, and HFE touchpoints.",
        "angles_zh": ["ISO 14971 的流程框架", "hazard、hazardous situation、harm 的关系", "标准如何与 use-related risk 协同"],
        "angles_en": ["The process structure of ISO 14971", "The relationship among hazard, hazardous situation, and harm", "How the standard interacts with use-related risk"],
        "source_files": ["ISO 14971 and Medical Device RM.pptx"],
        "cross_refs": ["medical_device_urra", "medical_device_use_errors", "epipen_workbook"],
    },
    {
        "slug": "medical_device_urra",
        "section": "HFE_Medical_Devices",
        "zh_nav": "医疗器械URRA",
        "en_nav": "Medical Device URRA",
        "zh_title": "医疗器械中的 URRA",
        "en_title": "URRA in Medical Devices",
        "zh_pitch": "本页聚焦医疗器械场景下的 URRA，把 use scenario、critical task、harm 路径与控制策略放到监管语境中理解。",
        "en_pitch": "This page focuses on URRA within medical devices, connecting use scenarios, critical tasks, harm pathways, and controls in a regulated context.",
        "angles_zh": ["医疗器械 URRA 与一般 URRA 有何不同", "critical task 如何界定", "控制措施如何回到界面、标签与 IFU"],
        "angles_en": ["How medical-device URRA differs from generic URRA", "How critical tasks are defined", "How controls return to interface, labeling, and IFU design"],
        "source_files": ["06 URRA in Medical Devices.pptx"],
        "cross_refs": ["iso_14971", "medical_device_use_errors", "epipen_workbook"],
    },
    {
        "slug": "medical_device_use_errors",
        "section": "HFE_Medical_Devices",
        "zh_nav": "医疗器械使用错误",
        "en_nav": "Medical Device Use Errors",
        "zh_title": "医疗器械中的使用错误",
        "en_title": "Use Errors in Medical Devices",
        "zh_pitch": "本页聚焦医疗器械中的 use error、异常使用与界面诱发风险，强调错误并非只来自操作者，而是来自整个使用系统。",
        "en_pitch": "This page focuses on use error, abnormal use, and interface-induced risk in medical devices, emphasizing that failure emerges from the use system rather than the operator alone.",
        "angles_zh": ["use error 与 misuse 的边界", "界面与说明如何诱发错误", "为什么 use error 要和 validation 结合看"],
        "angles_en": ["The boundary between use error and misuse", "How interfaces and instructions induce error", "Why use error must be tied to validation"],
        "source_files": ["07 Use Errors.pptx"],
        "cross_refs": ["iso_14971", "medical_device_urra", "task_analysis"],
    },
    {
        "slug": "epipen_workbook",
        "section": "HFE_Medical_Devices",
        "zh_nav": "EpiPen工作簿",
        "en_nav": "EpiPen Workbook",
        "zh_title": "EpiPen URRA 工作簿",
        "en_title": "The EpiPen URRA Workbook",
        "zh_pitch": "本页用 EpiPen 表格示例把任务分析、潜在 use error、hazard、harm、severity、critical task 与 mitigation 映射成可执行的工作底稿。",
        "en_pitch": "This page uses the EpiPen workbook to turn task analysis, potential use error, hazard, harm, severity, critical tasks, and mitigation into a concrete working artifact.",
        "angles_zh": ["表格字段如何对应风险链条", "工作簿如何支持团队协作", "从表格到验证任务如何落地"],
        "angles_en": ["How the table fields map to the risk chain", "How the workbook supports teamwork", "How workbook entries connect to validation tasks"],
        "source_files": ["10 URRA EpiPen.xlsx"],
        "cross_refs": ["task_analysis", "urra_methods", "medical_device_urra"],
    },
    {
        "slug": "aviation_automation_intro",
        "section": "HFE_Aviation_Automation",
        "zh_nav": "航空与自动化导论",
        "en_nav": "Aviation and Automation Intro",
        "zh_title": "航空与自动化导论",
        "en_title": "Introduction to Aviation and Automation",
        "zh_pitch": "本页用航空与自动化导论材料建立自动化收益、自动化陷阱与飞行员角色变化的总体框架。",
        "en_pitch": "This page establishes the high-level frame of automation benefits, automation traps, and the evolving role of the pilot.",
        "angles_zh": ["自动化为什么既能增效也能增险", "角色分配如何影响监控与干预", "航空系统为何是自动化 HFE 的经典实验场"],
        "angles_en": ["Why automation can both improve and increase risk", "How role allocation shapes monitoring and intervention", "Why aviation is a classic HFE laboratory for automation"],
        "source_files": ["Intro to Aviation and Automation 1-26-26-2.pdf"],
        "cross_refs": ["crew_resource_management", "displays_and_alerts", "automated_vehicles"],
    },
    {
        "slug": "crew_resource_management",
        "section": "HFE_Aviation_Automation",
        "zh_nav": "CRM",
        "en_nav": "CRM",
        "zh_title": "Crew Resource Management",
        "en_title": "Crew Resource Management",
        "zh_pitch": "本页整理 CRM 的沟通、协调、领导、监控与资源整合逻辑，把人因工程从个人表现扩展到团队表现。",
        "en_pitch": "This page summarizes communication, coordination, leadership, monitoring, and resource integration in CRM, extending HFE from individual to team performance.",
        "angles_zh": ["CRM 为什么是团队层面的人因能力", "沟通与授权如何影响安全裕度", "团队监控如何补偿个体局限"],
        "angles_en": ["Why CRM is a team-level HFE capability", "How communication and authority affect safety margins", "How team monitoring compensates for individual limits"],
        "source_files": ["CRM 1-28-2026.pdf"],
        "cross_refs": ["aviation_automation_intro", "attention_and_monitoring", "situation_awareness"],
    },
    {
        "slug": "displays_and_alerts",
        "section": "HFE_Aviation_Automation",
        "zh_nav": "显示与告警",
        "en_nav": "Displays and Alerts",
        "zh_title": "显示与告警",
        "en_title": "Displays and Alerts",
        "zh_pitch": "本页聚焦显示系统与告警设计，整理 salience、prioritization、false alarm 与 workload 之间的取舍。",
        "en_pitch": "This page focuses on display systems and alert design, organizing the tradeoffs among salience, prioritization, false alarms, and workload.",
        "angles_zh": ["告警为什么既要显著又不能淹没操作者", "显示层级如何支持快速诊断", "false alarm 与 trust erosion 的关系"],
        "angles_en": ["Why alerts must be salient without overwhelming the operator", "How display hierarchy supports rapid diagnosis", "How false alarms erode trust"],
        "source_files": ["Displays and Alerts 2-4-26.pdf"],
        "cross_refs": ["attention_and_monitoring", "checklists_and_procedures", "aviation_automation_intro"],
    },
    {
        "slug": "checklists_and_procedures",
        "section": "HFE_Aviation_Automation",
        "zh_nav": "检查单与程序",
        "en_nav": "Checklists and Procedures",
        "zh_title": "检查单与程序",
        "en_title": "Checklists and Procedures",
        "zh_pitch": "本页围绕 checklists 与 procedures 的设计原则，说明标准化如何减少记忆负担并支持团队协同。",
        "en_pitch": "This page explains the design logic of checklists and procedures and how standardization reduces memory burden while supporting team coordination.",
        "angles_zh": ["程序化为什么能提升可靠性", "checklist 设计如何兼顾速度与准确", "例外情形如何避免机械照本宣科"],
        "angles_en": ["Why proceduralization improves reliability", "How checklist design balances speed and accuracy", "How exceptions avoid rote compliance"],
        "source_files": ["Checklists and Procedures 3-23-26.pdf"],
        "cross_refs": ["crew_resource_management", "displays_and_alerts", "operational_risk"],
    },
    {
        "slug": "automated_vehicles",
        "section": "HFE_Aviation_Automation",
        "zh_nav": "自动化车辆",
        "en_nav": "Automated Vehicles",
        "zh_title": "自动化车辆",
        "en_title": "Automated Vehicles",
        "zh_pitch": "本页将航空自动化的经验延展到 automated vehicles，关注监控、handoff、trust calibration 与责任分配。",
        "en_pitch": "This page extends the automation discussion to automated vehicles, focusing on monitoring, handoff, trust calibration, and responsibility allocation.",
        "angles_zh": ["自动驾驶中的接管问题", "trust calibration 如何影响依赖与干预", "航空自动化经验能迁移哪些教训"],
        "angles_en": ["The takeover problem in vehicle automation", "How trust calibration shapes reliance and intervention", "Which lessons transfer from aviation automation"],
        "source_files": ["Automated Vehicles 3-11-26-2.pdf"],
        "cross_refs": ["aviation_automation_intro", "displays_and_alerts", "boeing_737max_and_ethics"],
    },
    {
        "slug": "attention_and_monitoring",
        "section": "HFE_Human_Performance",
        "zh_nav": "注意与监控",
        "en_nav": "Attention and Monitoring",
        "zh_title": "注意与监控",
        "en_title": "Attention and Monitoring",
        "zh_pitch": "本页整理注意资源、监控任务、automation complacency 与 vigilance decrement 等核心问题。",
        "en_pitch": "This page organizes attention resources, monitoring tasks, automation complacency, and vigilance decrement.",
        "angles_zh": ["注意为什么是有限资源", "监控任务为何特别容易衰退", "自动化如何改变人的注意分配"],
        "angles_en": ["Why attention is a limited resource", "Why monitoring tasks are vulnerable to decay", "How automation changes attentional allocation"],
        "source_files": ["Attention and Monitoring 2-9- 2026.pdf"],
        "cross_refs": ["displays_and_alerts", "fatigue_and_sleep", "situation_awareness"],
    },
    {
        "slug": "fatigue_and_sleep",
        "section": "HFE_Human_Performance",
        "zh_nav": "疲劳与睡眠",
        "en_nav": "Fatigue and Sleep",
        "zh_title": "疲劳与睡眠",
        "en_title": "Fatigue and Sleep",
        "zh_pitch": "本页围绕疲劳、生理节律与睡眠不足对表现的影响，整理人因工程里对生理约束的理解。",
        "en_pitch": "This page focuses on fatigue, circadian rhythm, and sleep loss as physiological constraints on performance.",
        "angles_zh": ["疲劳如何影响反应、判断与监控", "生理节律为什么会改变风险窗口", "管理疲劳不能只靠个人自律"],
        "angles_en": ["How fatigue affects response, judgment, and monitoring", "Why circadian rhythm changes the risk window", "Why fatigue management cannot rely on personal discipline alone"],
        "source_files": ["Sp26_Fatigue+Sleep_20260211.pdf"],
        "cross_refs": ["attention_and_monitoring", "stress_and_decision_making", "operational_risk"],
    },
    {
        "slug": "stress_and_decision_making",
        "section": "HFE_Human_Performance",
        "zh_nav": "压力与决策",
        "en_nav": "Stress and Decision Making",
        "zh_title": "压力、决策与分心",
        "en_title": "Stress, Decision Making, and Distraction",
        "zh_pitch": "本页将压力、决策、分心与认知偏差放在一起，说明高负荷场景下操作者为何更容易走向次优选择。",
        "en_pitch": "This page combines stress, decision making, distraction, and cognitive bias to explain why operators drift toward suboptimal choices under load.",
        "angles_zh": ["压力如何缩窄注意与策略空间", "分心如何侵蚀程序执行与监控", "决策偏差为何常与时间压力绑定"],
        "angles_en": ["How stress narrows attention and strategy space", "How distraction erodes procedure following and monitoring", "Why decision bias often couples with time pressure"],
        "source_files": ["Sp26_StressDecisionMaking_20260225.pdf", "Sp26_DecisionMakingDistraction_20260302.pdf"],
        "cross_refs": ["attention_and_monitoring", "fatigue_and_sleep", "boeing_737max_and_ethics"],
    },
    {
        "slug": "situation_awareness",
        "section": "HFE_Human_Performance",
        "zh_nav": "情境意识",
        "en_nav": "Situation Awareness",
        "zh_title": "情境意识",
        "en_title": "Situation Awareness",
        "zh_pitch": "本页围绕 perception、comprehension、projection 三层情境意识展开，连接信息获取与未来状态预判。",
        "en_pitch": "This page covers the perception, comprehension, and projection layers of situation awareness and links current information to future-state prediction.",
        "angles_zh": ["情境意识的三层结构", "为什么信息可见不等于真正理解", "预测未来状态为何决定有效干预"],
        "angles_en": ["The three-layer structure of situation awareness", "Why visible information does not equal understanding", "Why future-state projection shapes effective intervention"],
        "source_files": ["Sp26_SituationAwareness_20260325.pdf"],
        "cross_refs": ["attention_and_monitoring", "crew_resource_management", "spatial_disorientation"],
    },
    {
        "slug": "spatial_disorientation",
        "section": "HFE_Human_Performance",
        "zh_nav": "空间定向错觉",
        "en_nav": "Spatial Disorientation",
        "zh_title": "空间定向错觉",
        "en_title": "Spatial Disorientation",
        "zh_pitch": "本页整理空间定向错觉的生理基础、知觉陷阱与操作风险，强调感官系统与仪表信息冲突时的人因问题。",
        "en_pitch": "This page summarizes the physiology, perceptual traps, and operational risk of spatial disorientation when sensory systems conflict with instruments.",
        "angles_zh": ["感觉系统为什么会误导操作者", "仪表扫描如何对抗错觉", "空间定向问题为什么高度危险"],
        "angles_en": ["Why sensory systems can mislead operators", "How instrument scan counteracts illusion", "Why spatial disorientation is so hazardous"],
        "source_files": ["Sp26_SpDisorientation_20260330.pdf"],
        "cross_refs": ["situation_awareness", "attention_and_monitoring", "fatigue_and_sleep"],
    },
    {
        "slug": "operational_risk",
        "section": "HFE_Cases_Ethics",
        "zh_nav": "Operational Risk",
        "en_nav": "Operational Risk",
        "zh_title": "Operational Risk",
        "en_title": "Operational Risk",
        "zh_pitch": "本页围绕 operational risk 的构成因素、组织条件与现实运行约束，展示风险并不只属于单次任务，而是属于运行系统。",
        "en_pitch": "This page examines the ingredients of operational risk, organizational conditions, and real-world constraints to show that risk belongs to the operating system, not just a single task.",
        "angles_zh": ["运行风险为何超出单个错误事件", "组织约束如何塑造一线行为", "风险管理为何需要持续监测而不是一次性修补"],
        "angles_en": ["Why operational risk exceeds a single error event", "How organizational constraints shape frontline behavior", "Why risk management requires ongoing monitoring"],
        "source_files": ["Operational Risk 2-18-2026.pdf"],
        "cross_refs": ["swiss_cheese", "error_analysis_methods", "checklists_and_procedures"],
    },
    {
        "slug": "cardosi_case",
        "section": "HFE_Cases_Ethics",
        "zh_nav": "Cardosi案例",
        "en_nav": "Cardosi Case",
        "zh_title": "Cardosi 案例",
        "en_title": "The Cardosi Case",
        "zh_pitch": "本页围绕 Cardosi 材料提炼具体案例中的显示、程序、沟通与情境理解问题，作为把方法落到案例的中间层。",
        "en_pitch": "This page distills the Cardosi material into a concrete case where displays, procedures, communication, and situation understanding interact.",
        "angles_zh": ["案例中哪些设计与流程因素最关键", "如何从具体事件回看 HFE 原则", "案例分析如何反哺一般方法"],
        "angles_en": ["Which design and process factors matter most in the case", "How to read HFE principles back into a concrete event", "How case analysis feeds back into general methods"],
        "source_files": ["Cardosi 2-2-2026.pdf"],
        "cross_refs": ["displays_and_alerts", "crew_resource_management", "situation_awareness"],
    },
    {
        "slug": "f16_analysis_prompts",
        "section": "HFE_Cases_Ethics",
        "zh_nav": "F16分析提示",
        "en_nav": "F16 Analysis Prompts",
        "zh_title": "F16 分析提示",
        "en_title": "F16 Analysis Prompts",
        "zh_pitch": "本页将 F16 prompts 视为案例分析模板，帮助读者把课程概念转化为结构化提问与论证路径。",
        "en_pitch": "This page treats the F16 prompts as a case-analysis template that converts course concepts into structured questions and argument paths.",
        "angles_zh": ["分析 prompts 如何组织观察维度", "如何把概念转成案例提问", "结构化 prompts 为什么有助于团队复盘"],
        "angles_en": ["How prompts organize dimensions of observation", "How concepts become case questions", "Why structured prompts help team review"],
        "source_files": ["Sp26_F16-AnalysisPrompts_20260219.pdf"],
        "cross_refs": ["error_analysis_methods", "cardosi_case", "boeing_737max_and_ethics"],
    },
    {
        "slug": "boeing_737max_and_ethics",
        "section": "HFE_Cases_Ethics",
        "zh_nav": "Boeing 737 Max与伦理",
        "en_nav": "Boeing 737 Max and Ethics",
        "zh_title": "Boeing 737 Max 与伦理",
        "en_title": "Boeing 737 Max and Ethics",
        "zh_pitch": "本页把 Boeing 737 Max 作为技术、组织、监管与伦理问题交织的典型案例，说明 HFE 不能脱离治理与责任讨论。",
        "en_pitch": "This page treats Boeing 737 Max as a canonical case where technical, organizational, regulatory, and ethical issues are inseparable from HFE.",
        "angles_zh": ["设计决策如何变成系统性伦理问题", "组织与监管如何影响安全结果", "技术失配为何常伴随责任分配失衡"],
        "angles_en": ["How design decisions become systemic ethical issues", "How organization and regulation shape safety outcomes", "Why technical mismatch often accompanies accountability mismatch"],
        "source_files": ["Boeing 737Max and Ethics 2-23-26.pdf"],
        "cross_refs": ["automated_vehicles", "operational_risk", "stress_and_decision_making"],
    },
]

PAGE_BY_SLUG = {page["slug"]: page for page in PAGES}
SOURCE_TO_PAGE = {}
for page in PAGES:
    for source_file in page["source_files"]:
        if source_file in SOURCE_TO_PAGE:
            raise ValueError(f"Duplicate source mapping for {source_file}")
        SOURCE_TO_PAGE[source_file] = page["slug"]


def page_doc_rel(page_slug: str, lang: str = "zh") -> str:
    page = PAGE_BY_SLUG[page_slug]
    suffix = ".en.md" if lang == "en" else ".md"
    return f"{page['section']}/{page['slug']}{suffix}"


def rel_link(from_doc_rel: str, to_doc_rel: str) -> str:
    from_dir = posixpath.dirname(from_doc_rel)
    return posixpath.relpath(to_doc_rel, from_dir or ".")


def ensure_clean_generation_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    (DOCS_DIR / "assets").mkdir(parents=True, exist_ok=True)
    if ASSET_ROOT.exists():
        shutil.rmtree(ASSET_ROOT)
    ASSET_ROOT.mkdir(parents=True, exist_ok=True)
    if VISUAL_ROOT.exists():
        shutil.rmtree(VISUAL_ROOT)
    VISUAL_ROOT.mkdir(parents=True, exist_ok=True)
    for section_slug in SECTIONS:
        section_dir = DOCS_DIR / section_slug
        if section_dir.exists():
            shutil.rmtree(section_dir)
        section_dir.mkdir(parents=True, exist_ok=True)


def zip_entry_is_noise(name: str) -> bool:
    pure = PurePosixPath(name)
    return (
        name.endswith("/")
        or name.startswith("__MACOSX/")
        or pure.name in {".DS_Store"}
        or pure.name.startswith("._")
    )


def stable_source_slug(name: str) -> str:
    stem = Path(name).stem.lower()
    stem = re.sub(r"[^a-z0-9]+", "-", stem).strip("-")
    return stem or "source"


def sanitize_filename_fragment(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "-", value).strip("-") or "asset"


def write_visual_asset(source_file: str, logical_name: str, payload: bytes) -> str:
    rel_path = PurePosixPath("assets/visuals") / stable_source_slug(source_file) / sanitize_filename_fragment(logical_name)
    abs_path = DOCS_DIR / rel_path
    abs_path.parent.mkdir(parents=True, exist_ok=True)
    abs_path.write_bytes(payload)
    return rel_path.as_posix()


def normalize_text(text: str) -> str:
    text = text.replace("\xa0", " ").replace("\u200b", "")
    text = text.replace("```", "'''")
    return text.strip()


def split_text_lines(value: str) -> list[str]:
    lines = [normalize_text(line) for line in str(value).splitlines()]
    return [line for line in lines if line]


def png_dimensions(payload: bytes) -> tuple[int | None, int | None]:
    if len(payload) < 24 or payload[:8] != b"\x89PNG\r\n\x1a\n":
        return None, None
    return int.from_bytes(payload[16:20], "big"), int.from_bytes(payload[20:24], "big")


def gif_dimensions(payload: bytes) -> tuple[int | None, int | None]:
    if len(payload) < 10 or payload[:3] != b"GIF":
        return None, None
    return int.from_bytes(payload[6:8], "little"), int.from_bytes(payload[8:10], "little")


def jpeg_dimensions(payload: bytes) -> tuple[int | None, int | None]:
    if len(payload) < 4 or payload[:2] != b"\xff\xd8":
        return None, None
    index = 2
    while index + 9 < len(payload):
        if payload[index] != 0xFF:
            index += 1
            continue
        marker = payload[index + 1]
        index += 2
        if marker in {0xD8, 0xD9}:
            continue
        if index + 2 > len(payload):
            break
        segment_len = int.from_bytes(payload[index:index + 2], "big")
        if segment_len < 2 or index + segment_len > len(payload):
            break
        if marker in {0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7, 0xC9, 0xCA, 0xCB, 0xCD, 0xCE, 0xCF}:
            if index + 7 <= len(payload):
                height = int.from_bytes(payload[index + 3:index + 5], "big")
                width = int.from_bytes(payload[index + 5:index + 7], "big")
                return width, height
            break
        index += segment_len
    return None, None


def image_dimensions(payload: bytes, ext: str) -> tuple[int | None, int | None]:
    if ext == ".png":
        return png_dimensions(payload)
    if ext == ".gif":
        return gif_dimensions(payload)
    if ext in {".jpg", ".jpeg"}:
        return jpeg_dimensions(payload)
    return None, None


def is_meaningful_visual(payload: bytes, ext: str) -> bool:
    if ext == ".svg":
        return len(payload) >= 600
    if len(payload) < 3500:
        return False
    width, height = image_dimensions(payload, ext)
    if width and height and width < 140 and height < 140 and len(payload) < 50000:
        return False
    return True


def resolve_zip_target(base_name: str, target: str) -> str:
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_name), target))


def extract_ppt_paragraphs(xml_bytes: bytes) -> list[str]:
    root = ET.fromstring(xml_bytes)
    extracted: list[str] = []
    for para in root.findall(".//a:p", PPT_NS):
        text = "".join(t.text or "" for t in para.findall(".//a:t", PPT_NS))
        extracted.extend(split_text_lines(text))
    return extracted


def build_unit(
    source_file: str,
    source_type: str,
    locator: str,
    raw_text: str,
    topic_slug: str,
    unit_index: int,
    asset_rel_path: str,
) -> dict:
    unit_id = f"{stable_source_slug(source_file)}-{unit_index:04d}"
    return {
        "unit_id": unit_id,
        "source_file": source_file,
        "source_type": source_type,
        "locator": locator,
        "language": "en",
        "raw_text": raw_text,
        "topic_slug": topic_slug,
        "asset_rel_path": asset_rel_path,
    }


def extract_pptx_units(payload: bytes, source_file: str, topic_slug: str, asset_rel_path: str) -> tuple[list[dict], list[str]]:
    units: list[dict] = []
    warnings: list[str] = []
    counter = 1
    with zipfile.ZipFile(io.BytesIO(payload)) as inner:
        slide_names = sorted(
            [name for name in inner.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)],
            key=lambda value: int(re.search(r"slide(\d+)\.xml", value).group(1)),
        )
        note_names = sorted(
            [name for name in inner.namelist() if re.fullmatch(r"ppt/notesSlides/notesSlide\d+\.xml", name)],
            key=lambda value: int(re.search(r"notesSlide(\d+)\.xml", value).group(1)),
        )
        for slide_name in slide_names:
            slide_no = int(re.search(r"slide(\d+)\.xml", slide_name).group(1))
            for idx, line in enumerate(extract_ppt_paragraphs(inner.read(slide_name)), start=1):
                units.append(build_unit(source_file, "pptx", f"slide:{slide_no}:p:{idx}", line, topic_slug, counter, asset_rel_path))
                counter += 1
        for note_name in note_names:
            note_no = int(re.search(r"notesSlide(\d+)\.xml", note_name).group(1))
            note_lines = extract_ppt_paragraphs(inner.read(note_name))
            if not note_lines:
                warnings.append(f"{source_file}: notesSlide {note_no} has no extracted text")
            for idx, line in enumerate(note_lines, start=1):
                units.append(build_unit(source_file, "pptx", f"notes:{note_no}:p:{idx}", line, topic_slug, counter, asset_rel_path))
                counter += 1
    return units, warnings


def extract_pdf_units(payload: bytes, source_file: str, topic_slug: str, asset_rel_path: str) -> tuple[list[dict], list[str]]:
    units: list[dict] = []
    warnings: list[str] = []
    counter = 1
    with fitz.open(stream=payload, filetype="pdf") as document:
        for page_index in range(document.page_count):
            page = document.load_page(page_index)
            lines = split_text_lines(page.get_text("text"))
            if not lines:
                warnings.append(f"{source_file}: page {page_index + 1} produced no text")
            for line_index, line in enumerate(lines, start=1):
                units.append(build_unit(source_file, "pdf", f"page:{page_index + 1}:line:{line_index}", line, topic_slug, counter, asset_rel_path))
                counter += 1
    return units, warnings


def extract_xlsx_units(payload: bytes, source_file: str, topic_slug: str, asset_rel_path: str) -> tuple[list[dict], list[str]]:
    units: list[dict] = []
    warnings: list[str] = []
    counter = 1
    workbook = load_workbook(io.BytesIO(payload), data_only=True)
    for worksheet in workbook.worksheets:
        seen_text = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_lines = split_text_lines(str(cell.value))
                for line_index, line in enumerate(cell_lines, start=1):
                    locator = f"sheet:{worksheet.title}!{cell.coordinate}:line:{line_index}"
                    units.append(build_unit(source_file, "xlsx", locator, line, topic_slug, counter, asset_rel_path))
                    counter += 1
                    seen_text += 1
        if seen_text == 0:
            warnings.append(f"{source_file}: worksheet '{worksheet.title}' has no extracted text")
    return units, warnings


def extract_pptx_visuals(payload: bytes, source_file: str) -> tuple[list[dict], list[str]]:
    visuals: list[dict] = []
    warnings: list[str] = []
    with zipfile.ZipFile(io.BytesIO(payload)) as inner:
        inner_names = set(inner.namelist())
        slide_names = sorted(
            [name for name in inner_names if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)],
            key=lambda value: int(re.search(r"slide(\d+)\.xml", value).group(1)),
        )
        for slide_name in slide_names:
            slide_no = int(re.search(r"slide(\d+)\.xml", slide_name).group(1))
            rel_name = f"ppt/slides/_rels/slide{slide_no}.xml.rels"
            if rel_name not in inner_names:
                continue
            slide_root = ET.fromstring(inner.read(slide_name))
            rel_root = ET.fromstring(inner.read(rel_name))
            rel_targets = {
                rel.attrib.get("Id"): rel.attrib.get("Target")
                for rel in rel_root.findall(".//rel:Relationship", PKG_REL_NS)
                if rel.attrib.get("Id") and rel.attrib.get("Target")
            }
            seen_targets: set[str] = set()
            visual_index = 1
            for blip in slide_root.findall(".//a:blip", PPT_NS):
                rel_id = blip.attrib.get(f"{{{REL_NS_URI}}}embed")
                if not rel_id or rel_id not in rel_targets:
                    continue
                resolved_target = resolve_zip_target(slide_name, rel_targets[rel_id])
                if resolved_target in seen_targets:
                    continue
                seen_targets.add(resolved_target)
                ext = Path(resolved_target).suffix.lower()
                if ext not in SUPPORTED_VISUAL_EXTENSIONS:
                    continue
                try:
                    asset_bytes = inner.read(resolved_target)
                except KeyError:
                    warnings.append(f"{source_file}: missing visual target {resolved_target}")
                    continue
                if not is_meaningful_visual(asset_bytes, ext):
                    continue
                asset_rel_path = write_visual_asset(
                    source_file,
                    f"slide-{slide_no:02d}-{Path(resolved_target).name}",
                    asset_bytes,
                )
                visuals.append(
                    {
                        "source_file": source_file,
                        "source_type": "pptx",
                        "locator": f"slide:{slide_no}:image:{visual_index}",
                        "asset_rel_path": asset_rel_path,
                    }
                )
                visual_index += 1
    return visuals, warnings


def select_pdf_preview_pages(document: fitz.Document) -> list[int]:
    selected: list[int] = []
    if document.page_count == 0:
        return selected
    selected.append(1)

    image_pages: list[int] = []
    low_text_pages: list[tuple[int, int]] = []
    for page_no in range(1, document.page_count + 1):
        page = document.load_page(page_no - 1)
        text_line_count = len(split_text_lines(page.get_text("text")))
        image_count = len(page.get_images(full=True))
        if image_count > 0 and page_no not in image_pages:
            image_pages.append(page_no)
        low_text_pages.append((text_line_count, page_no))

    for page_no in image_pages:
        if page_no not in selected:
            selected.append(page_no)
        if len(selected) >= 3:
            return sorted(selected)

    for _, page_no in sorted(low_text_pages):
        if page_no not in selected:
            selected.append(page_no)
        if len(selected) >= 3:
            break
    return sorted(selected)


def extract_pdf_visuals(payload: bytes, source_file: str) -> tuple[list[dict], list[str]]:
    visuals: list[dict] = []
    warnings: list[str] = []
    with fitz.open(stream=payload, filetype="pdf") as document:
        for page_no in select_pdf_preview_pages(document):
            page = document.load_page(page_no - 1)
            try:
                pixmap = page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5), alpha=False)
                png_bytes = pixmap.tobytes("png")
            except Exception as exc:  # pragma: no cover - defensive against malformed PDFs
                warnings.append(f"{source_file}: failed to render page preview {page_no} ({exc})")
                continue
            asset_rel_path = write_visual_asset(source_file, f"page-{page_no:02d}.png", png_bytes)
            visuals.append(
                {
                    "source_file": source_file,
                    "source_type": "pdf",
                    "locator": f"page:{page_no}:preview",
                    "asset_rel_path": asset_rel_path,
                }
            )
    return visuals, warnings


def extract_xlsx_visuals(payload: bytes, source_file: str) -> tuple[list[dict], list[str]]:
    del payload, source_file
    return [], []


def extractor_for(source_type: str):
    if source_type == "pptx":
        return extract_pptx_units
    if source_type == "pdf":
        return extract_pdf_units
    if source_type == "xlsx":
        return extract_xlsx_units
    raise ValueError(f"Unsupported source type: {source_type}")


def visual_extractor_for(source_type: str):
    if source_type == "pptx":
        return extract_pptx_visuals
    if source_type == "pdf":
        return extract_pdf_visuals
    if source_type == "xlsx":
        return extract_xlsx_visuals
    raise ValueError(f"Unsupported source type: {source_type}")


def extract_all_sources() -> tuple[list[dict], dict[str, dict], list[str], dict]:
    source_units: list[dict] = []
    manifest: dict[str, dict] = {}
    warnings: list[str] = []
    counts = Counter()
    total_visuals = 0

    for archive in SOURCE_ARCHIVES:
        zip_path = WORKSPACE_ROOT / archive["zip_name"]
        if not zip_path.exists():
            raise FileNotFoundError(f"Missing source archive: {zip_path}")
        with zipfile.ZipFile(zip_path) as outer:
            for info in outer.infolist():
                if zip_entry_is_noise(info.filename):
                    continue
                file_name = PurePosixPath(info.filename).name
                if file_name not in SOURCE_TO_PAGE:
                    raise ValueError(f"Unmapped source file: {file_name}")
                payload = outer.read(info.filename)
                source_type = Path(file_name).suffix.lower().lstrip(".")
                topic_slug = SOURCE_TO_PAGE[file_name]
                asset_rel = PurePosixPath("assets/source_files") / archive["asset_dir"] / file_name
                asset_abs = DOCS_DIR / asset_rel
                asset_abs.parent.mkdir(parents=True, exist_ok=True)
                asset_abs.write_bytes(payload)
                extractor = extractor_for(source_type)
                units, source_warnings = extractor(payload, file_name, topic_slug, asset_rel.as_posix())
                visual_extractor = visual_extractor_for(source_type)
                visuals, visual_warnings = visual_extractor(payload, file_name)
                warnings.extend(source_warnings)
                warnings.extend(visual_warnings)
                source_units.extend(units)
                counts[source_type] += 1
                total_visuals += len(visuals)
                manifest[file_name] = {
                    "source_file": file_name,
                    "source_type": source_type,
                    "topic_slug": topic_slug,
                    "archive_zip": archive["zip_name"],
                    "archive_entry": info.filename,
                    "asset_rel_path": asset_rel.as_posix(),
                    "unit_count": len(units),
                    "visual_count": len(visuals),
                    "visuals": visuals,
                }

    expected_counts = {"pptx": 10, "pdf": 19, "xlsx": 1}
    if dict(counts) != expected_counts:
        raise ValueError(f"Unexpected source counts: {dict(counts)} != {expected_counts}")

    summary = {
        "expected_counts": expected_counts,
        "observed_counts": dict(counts),
        "total_sources": len(manifest),
        "total_units": len(source_units),
        "total_visuals": total_visuals,
        "warnings": warnings,
    }
    return source_units, manifest, warnings, summary


def select_highlights(units: list[dict], limit: int = 8) -> list[str]:
    highlights: list[str] = []
    seen: set[str] = set()
    for unit in units:
        cleaned = re.sub(r"\s+", " ", unit["raw_text"]).strip()
        if len(cleaned) < 18:
            continue
        if cleaned.lower() in seen:
            continue
        if re.fullmatch(r"[\d\W]+", cleaned):
            continue
        highlights.append(cleaned)
        seen.add(cleaned.lower())
        if len(highlights) >= limit:
            break
    return highlights


def format_source_table(page: dict, manifest: dict[str, dict], current_doc_rel: str) -> list[str]:
    lines = [
        "| Source | Type | Text Units | Visuals | Download |",
        "| --- | --- | ---: | ---: | --- |",
    ]
    for source_file in page["source_files"]:
        source_meta = manifest[source_file]
        download_link = rel_link(current_doc_rel, source_meta["asset_rel_path"])
        lines.append(
            f"| `{source_file}` | `{source_meta['source_type']}` | {source_meta['unit_count']} | "
            f"{source_meta.get('visual_count', 0)} | [open]({download_link}) |"
        )
    return lines


def format_cross_refs(page: dict, current_doc_rel: str, lang: str) -> list[str]:
    lines: list[str] = []
    for target_slug in page["cross_refs"]:
        target_page = PAGE_BY_SLUG[target_slug]
        target_rel = page_doc_rel(target_slug, lang)
        label = target_page["zh_title"] if lang == "zh" else target_page["en_title"]
        lines.append(f"- [{label}]({rel_link(current_doc_rel, target_rel)})")
    return lines


def teaching_content_for(page_slug: str, lang: str) -> str:
    content = PAGE_CONTENT.get(page_slug, {})
    if lang in content:
        return content[lang]
    if lang == "zh":
        return "## 一眼看懂\n\n这一页的知识点讲解仍在补充中，当前先保留了来源、图示和完整原文，便于继续深化整理。"
    return "## At a Glance\n\nThis page is still waiting for a fuller teaching draft. The sources, visuals, and full transcription remain available below."


def normalize_teaching_content(content: str, lang: str) -> str:
    if lang == "zh":
        return content.replace("## 一眼看懂", "## 核心概念", 1)
    return content.replace("## At a Glance", "## Core Idea", 1)


def render_logic_outline(page: dict, lang: str) -> list[str]:
    title = "## 这页的逻辑顺序" if lang == "zh" else "## Reading Logic"
    intro = (
        "建议按下面的顺序读这页，这样会更像老师在课堂上带着你展开概念。"
        if lang == "zh"
        else "Read the page in this order to follow the lecture logic rather than treating it as a flat summary."
    )
    angles = page["angles_zh"] if lang == "zh" else page["angles_en"]
    lines = [title, "", intro, ""]
    for index, angle in enumerate(angles, start=1):
        lines.append(f"{index}. {angle}")
    return lines


def locator_outline_key(locator: str) -> tuple[str, int, int] | None:
    slide_match = re.match(r"slide:(\d+):p:(\d+)$", locator)
    if slide_match:
        return "slide", int(slide_match.group(1)), int(slide_match.group(2))
    page_match = re.match(r"page:(\d+):line:(\d+)$", locator)
    if page_match:
        return "page", int(page_match.group(1)), int(page_match.group(2))
    return None


def is_noise_outline_text(text: str) -> bool:
    stripped = normalize_text(text)
    if not stripped:
        return True
    if len(stripped) < 3 or len(stripped) > 110:
        return True
    if "http://" in stripped.lower() or "https://" in stripped.lower():
        return True
    if re.fullmatch(r"[\d\s/:\-().]+", stripped):
        return True
    if re.fullmatch(r"ENP-\d+", stripped, re.IGNORECASE):
        return True
    if re.fullmatch(r"page \d+", stripped, re.IGNORECASE):
        return True
    return False


def extract_source_outline(source_units: list[dict]) -> list[str]:
    grouped: dict[tuple[str, int], list[tuple[int, str]]] = defaultdict(list)
    for unit in source_units:
        outline_key = locator_outline_key(unit["locator"])
        if not outline_key:
            continue
        kind, block_no, line_no = outline_key
        grouped[(kind, block_no)].append((line_no, unit["raw_text"]))

    outline: list[str] = []
    seen: set[str] = set()
    for _, items in sorted(grouped.items(), key=lambda item: item[0][1]):
        candidates: list[str] = []
        for line_no, raw_text in sorted(items, key=lambda item: item[0]):
            if line_no > 4:
                break
            if is_noise_outline_text(raw_text):
                continue
            cleaned = re.sub(r"^[•\-]+\s*", "", normalize_text(raw_text))
            if cleaned:
                candidates.append(cleaned)
            if len(candidates) >= 2:
                break
        if not candidates:
            continue
        heading = " ".join(candidates[:2]) if sum(len(item) for item in candidates[:2]) <= 90 else candidates[0]
        heading = re.sub(r"\s+", " ", heading).strip(" -")
        normalized = heading.lower()
        if normalized in seen:
            continue
        seen.add(normalized)
        outline.append(heading)
        if len(outline) >= 8:
            break
    return outline


def render_classroom_outline(page: dict, units_by_source: dict[str, list[dict]], lang: str) -> list[str]:
    title = "## 课件里的讲解顺序" if lang == "zh" else "## Lecture Flow in the Source Material"
    intro = (
        "这一部分不是我主观概括，而是根据 PPT / PDF 的标题行和页首内容还原老师大致的课堂展开顺序。"
        if lang == "zh"
        else "This section reconstructs the lecture flow from slide titles and page-leading text instead of relying only on a hand-written summary."
    )
    lines = [title, "", intro, ""]
    has_outline = False
    for source_file in page["source_files"]:
        outline = extract_source_outline(units_by_source[source_file])
        if not outline:
            continue
        has_outline = True
        lines.append(f"### {source_file}")
        lines.append("")
        for index, item in enumerate(outline, start=1):
            lines.append(f"{index}. {item}")
        lines.append("")
    if not has_outline:
        lines.append(
            "当前没有从源文件中稳定提取出课件标题顺序，但下方仍保留了完整原文和源文件。"
            if lang == "zh"
            else "No stable slide/page outline was extracted for this topic, but the full source transcription remains below."
        )
    return lines


def format_visual_caption(visual: dict, lang: str) -> str:
    source_file = visual["source_file"]
    locator = visual["locator"]
    if locator.startswith("slide:"):
        slide_no = locator.split(":")[1]
        return f"{source_file} · 第 {slide_no} 张幻灯片" if lang == "zh" else f"{source_file} · slide {slide_no}"
    if locator.startswith("page:"):
        page_no = locator.split(":")[1]
        return f"{source_file} · 第 {page_no} 页预览" if lang == "zh" else f"{source_file} · page {page_no} preview"
    return source_file


def absolute_asset_url(asset_rel_path: str) -> str:
    return SITE_URL.rstrip("/") + "/" + quote(asset_rel_path.lstrip("/"), safe="/-_.()")


def render_visual_gallery(page: dict, manifest: dict[str, dict], current_doc_rel: str, lang: str) -> list[str]:
    title = "## 课件图示与页面预览" if lang == "zh" else "## Slide Figures and Page Previews"
    intro = (
        "下面展示从原始 PPT/PDF 中自动提取出的配图或页面预览。它们不是装饰图，而是正文讲解时应该对照着看的课堂材料。"
        if lang == "zh"
        else "These figures and page previews are extracted from the source slides/PDFs and are meant to be read together with the note content."
    )
    visuals: list[dict] = []
    for source_file in page["source_files"]:
        visuals.extend(manifest[source_file].get("visuals", []))

    lines = [title, "", intro, ""]
    if not visuals:
        lines.append(
            "本页源文件没有提取到适合展示的配图资产，但原始文件和逐行转写仍完整保留在本页底部。"
            if lang == "zh"
            else "No displayable visual assets were extracted for this page, but the raw files and full transcription remain available below."
        )
        return lines

    lines.append('<div class="note-visual-grid">')
    for visual in visuals:
        asset_href = absolute_asset_url(visual["asset_rel_path"])
        caption = format_visual_caption(visual, lang)
        lines.extend(
            [
                '  <figure class="note-visual">',
                f'    <img src="{html.escape(asset_href)}" alt="{html.escape(caption)}" loading="lazy">',
                f'    <figcaption>{html.escape(caption)}</figcaption>',
                "  </figure>",
            ]
        )
    lines.append("</div>")
    return lines


def render_transcript_blocks(
    source_files: list[str],
    units_by_source: dict[str, list[dict]],
    manifest: dict[str, dict],
    current_doc_rel: str,
    lang: str,
) -> list[str]:
    title_prefix = "原文转写与来源映射" if lang == "zh" else "Original Transcription and Coverage Mapping"
    download_prefix = "下载原件" if lang == "zh" else "Download source"
    mapped_prefix = "映射页面" if lang == "zh" else "Mapped page"
    intro = (
        "下面的折叠区块保留逐页/逐幻灯/逐单元原文。每一行前面的 `unit_id` 都能在 `data/coverage_map.json` 中找到对应页面映射。"
        if lang == "zh"
        else "The collapsible blocks below preserve page/slide-level source transcription. Each `unit_id` maps one-to-one in `data/coverage_map.json`."
    )
    lines = [f"## {title_prefix}", "", intro, ""]
    for source_file in source_files:
        source_meta = manifest[source_file]
        source_units = units_by_source[source_file]
        download_link = rel_link(current_doc_rel, source_meta["asset_rel_path"])
        caption = f"{source_file} | {len(source_units)} text units"
        lines.append(f'??? info "{caption}"')
        lines.append(f"    {download_prefix}: [{source_file}]({download_link})")
        lines.append(f"    {mapped_prefix}: `{source_meta['topic_slug']}`")
        lines.append("    ")
        lines.append("    ```text")
        for unit in source_units:
            raw_text = unit["raw_text"].replace("```", "'''")
            lines.append(f"    [{unit['unit_id']}] {unit['locator']} | {raw_text}")
        lines.append("    ```")
        lines.append("")
    return lines


def render_page(
    page: dict,
    manifest: dict[str, dict],
    units_by_page: dict[str, list[dict]],
    units_by_source: dict[str, list[dict]],
    lang: str,
) -> str:
    current_doc_rel = page_doc_rel(page["slug"], lang)
    section = SECTIONS[page["section"]]
    content_block = normalize_teaching_content(teaching_content_for(page["slug"], lang), lang)

    if lang == "zh":
        title = page["zh_title"]
        pitch = page["zh_pitch"]
        section_title = section["zh_title"]
        headers = {
            "sources": "## 资料范围与相关主题",
            "related": "## 相关主题",
        }
        source_intro = "正文先把知识点讲清楚；这里列出本页用到的原始文件，页尾折叠区块则保留完整逐行转写，便于你核对。"
    else:
        title = page["en_title"]
        pitch = page["en_pitch"]
        section_title = section["en_title"]
        headers = {
            "sources": "## Source Scope and Related Topics",
            "related": "## Related Topics",
        }
        source_intro = "The teaching notes come first. This section lists the source files used on the page, and the appendix keeps the full line-by-line transcription for verification."

    lines = [
        f"# {title}",
        "",
        pitch,
        "",
    ]
    lines.extend(render_classroom_outline(page, units_by_source, lang))
    lines.extend([""])
    lines.extend(render_logic_outline(page, lang))
    lines.extend(
        [
            "",
        content_block,
        "",
        ]
    )
    lines.extend(render_visual_gallery(page, manifest, current_doc_rel, lang))
    lines.extend(["", headers["sources"], "", source_intro, ""])
    lines.extend(
        [
            f"- {'所属分区' if lang == 'zh' else 'Section'}: `{section_title}`",
            f"- {'关联源文件数' if lang == 'zh' else 'Source files'}: {len(page['source_files'])}",
            f"- {'文本单元数' if lang == 'zh' else 'Text units'}: {len(units_by_page[page['slug']])}",
            f"- {'配图/预览数' if lang == 'zh' else 'Visuals/previews'}: {sum(manifest[source]['visual_count'] for source in page['source_files'])}",
            "",
        ]
    )
    lines.extend(format_source_table(page, manifest, current_doc_rel))
    lines.extend(["", headers["related"], ""])
    lines.extend(format_cross_refs(page, current_doc_rel, lang))
    lines.extend([""])
    lines.extend(render_transcript_blocks(page["source_files"], units_by_source, manifest, current_doc_rel, lang))
    return "\n".join(lines).rstrip() + "\n"


def render_section_index(section_slug: str, pages: list[dict], manifest: dict[str, dict], lang: str) -> str:
    section = SECTIONS[section_slug]
    current_doc_rel = f"{section_slug}/index{'.en' if lang == 'en' else ''}.md"
    if lang == "zh":
        title = section["zh_title"]
        intro = section["zh_intro"]
        page_label = "## 主题页面"
        source_label = "## 覆盖源文件"
        make_line = lambda page: f"- [{page['zh_title']}]({rel_link(current_doc_rel, page_doc_rel(page['slug'], 'zh'))}): {page['zh_pitch']}"
    else:
        title = section["en_title"]
        intro = section["en_intro"]
        page_label = "## Topic Pages"
        source_label = "## Covered Source Files"
        make_line = lambda page: f"- [{page['en_title']}]({rel_link(current_doc_rel, page_doc_rel(page['slug'], 'en'))}): {page['en_pitch']}"

    lines = [f"# {title}", "", intro, "", page_label, ""]
    lines.extend(make_line(page) for page in pages)
    lines.extend(["", source_label, ""])
    for page in pages:
        for source_file in page["source_files"]:
            source_meta = manifest[source_file]
            source_link = rel_link(current_doc_rel, source_meta["asset_rel_path"])
            lines.append(
                f"- `{source_file}` ({source_meta['source_type']}, {source_meta['unit_count']} units, "
                f"{source_meta.get('visual_count', 0)} visuals): [open]({source_link})"
            )
    return "\n".join(lines).rstrip() + "\n"


def render_home_page(lang: str) -> str:
    if lang == "zh":
        title = "人因工程课程笔记"
        intro = "这是一个围绕人因工程与 use-related risk 的双语 MkDocs 笔记站。内容按知识点主题组织，每一页都先讲知识点、再展示可提取图示，最后保留逐页/逐幻灯/逐单元的完整英文原文，确保每一行材料都能被追溯。"
        card_cta = "进入 →"
        coverage = [
            "- 源资料总数：`10 PPTX + 19 PDF + 1 XLSX = 30` 个文件",
            "- 组织方式：按知识点主题整理，不按原压缩包平铺",
            "- 页面结构：先读讲解，再看配图/预览，最后用逐行原文核对细节",
            "- 完整性机制：每个文本单元都写入 `data/source_units.jsonl`，并通过 `data/coverage_map.json` 一对一映射到页面底部折叠区块",
        ]
    else:
        title = "HFE Course Notes"
        intro = "This bilingual MkDocs site organizes human factors engineering and use-related risk materials by concept rather than raw archive. Each page now teaches the concept first, shows extracted visuals or page previews next, and preserves the full English source transcription afterward."
        card_cta = "Open →"
        coverage = [
            "- Source corpus: `10 PPTX + 19 PDF + 1 XLSX = 30` files",
            "- Organization: concept-first knowledge pages rather than flat archive listings",
            "- Page flow: explanation first, visuals/previews second, full source traceability last",
            "- Completeness: every text unit is written to `data/source_units.jsonl` and mapped one-to-one in `data/coverage_map.json`",
        ]

    lines = [
        "---",
        "hide:",
        "  - navigation",
        "  - toc",
        "---",
        "",
        f"# {title}",
        "",
        intro,
        "",
        "---",
        "",
        '<div class="grid cards" markdown>',
        "",
    ]
    for section_slug, section in SECTIONS.items():
        heading = section["zh_title"] if lang == "zh" else section["en_title"]
        body = section["zh_card"] if lang == "zh" else section["en_card"]
        href = f"{section_slug}/index.md" if lang == "zh" else f"{section_slug}/index.en.md"
        lines.extend(
            [
                f"-   **{heading}**",
                "",
                "    ---",
                "",
                f"    {body}",
                "",
                f"    [{card_cta}]({href}){{ .md-button }}",
                "",
            ]
        )
    lines.extend(["</div>", "", "---", ""])
    lines.extend(coverage)
    return "\n".join(lines).rstrip() + "\n"


def build_mkdocs_yml() -> str:
    translations = OrderedDict()
    translations["首页"] = "Home"
    for section in SECTIONS.values():
        translations[section["zh_nav"]] = section["en_nav"]
        translations[section["zh_index_nav"]] = section["en_index_nav"]
    for page in PAGES:
        translations[page["zh_nav"]] = page["en_nav"]

    lines = [
        "site_name: 人因工程课程笔记",
        "site_author: Jeff Liu",
        "site_url: https://jeffliulab.github.io/hfe-notes/",
        "",
        "theme:",
        "  name: material",
        "  custom_dir: overrides",
        "  features:",
        "    - navigation.tabs",
        "    - navigation.tabs.sticky",
        "    - navigation.indexes",
        "    - navigation.top",
        "    - navigation.path",
        "    - search.suggest",
        "    - search.highlight",
        "",
        "plugins:",
        "  - search",
        "  - i18n:",
        "      docs_structure: suffix",
        "      languages:",
        "        - locale: zh",
        "          default: true",
        "          name: 中文",
        "          build: true",
        "        - locale: en",
        "          name: English",
        "          build: true",
        "          site_name: HFE Course Notes",
        "          nav_translations:",
    ]
    for zh_label, en_label in translations.items():
        lines.append(f"            {zh_label}: {json.dumps(en_label, ensure_ascii=False)}")

    lines.extend(
        [
            "",
            "markdown_extensions:",
            "  - attr_list",
            "  - md_in_html",
            "  - admonition",
            "  - pymdownx.details",
            "  - pymdownx.superfences",
            "  - pymdownx.arithmatex:",
            "      generic: true",
            "  - mdx_truly_sane_lists:",
            "      nested_indent: 2",
            "",
            "extra_javascript:",
            "  - site_extra_styles/javascripts/katex.js",
            "  - site_extra_styles/javascripts/tab_dropdown.js",
            "  - https://unpkg.com/katex@0/dist/katex.min.js",
            "  - https://unpkg.com/katex@0/dist/contrib/auto-render.min.js",
            "",
            "extra_css:",
            "  - https://unpkg.com/katex@0/dist/katex.min.css",
            "  - site_extra_styles/stylesheets/extra.css",
            "  - site_extra_styles/stylesheets/correct_bullet_point.css",
            "",
            "nav:",
            "  - 首页: index.md",
        ]
    )

    pages_by_section: dict[str, list[dict]] = defaultdict(list)
    for page in PAGES:
        pages_by_section[page["section"]].append(page)

    for section_slug, section in SECTIONS.items():
        lines.append(f"  - {section['zh_nav']}:")
        lines.append(f"    - {section['zh_index_nav']}: {section_slug}/index.md")
        for page in pages_by_section[section_slug]:
            lines.append(f"    - {page['zh_nav']}: {section_slug}/{page['slug']}.md")
    return "\n".join(lines).rstrip() + "\n"


def write_file(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def build_coverage_map(source_units: list[dict]) -> dict:
    assignments = {}
    duplicates: list[str] = []
    for unit in source_units:
        if unit["unit_id"] in assignments:
            duplicates.append(unit["unit_id"])
        assignments[unit["unit_id"]] = {
            "topic_slug": unit["topic_slug"],
            "doc_path_zh": page_doc_rel(unit["topic_slug"], "zh"),
            "doc_path_en": page_doc_rel(unit["topic_slug"], "en"),
            "appendix_block": unit["source_file"],
            "locator": unit["locator"],
        }
    coverage = {
        "summary": {
            "total_units": len(source_units),
            "assigned_units": len(assignments),
            "duplicate_unit_ids": duplicates,
            "uncovered_units": [],
        },
        "unit_assignments": assignments,
    }
    if len(assignments) != len(source_units):
        raise ValueError("Coverage map lost unit assignments")
    return coverage


def build_docs(source_units: list[dict], manifest: dict[str, dict], summary: dict) -> None:
    pages_by_section: dict[str, list[dict]] = defaultdict(list)
    units_by_page: dict[str, list[dict]] = defaultdict(list)
    units_by_source: dict[str, list[dict]] = defaultdict(list)

    for page in PAGES:
        pages_by_section[page["section"]].append(page)
    for unit in source_units:
        units_by_page[unit["topic_slug"]].append(unit)
        units_by_source[unit["source_file"]].append(unit)

    write_file(REPO_ROOT / "mkdocs.yml", build_mkdocs_yml())
    write_file(DOCS_DIR / "index.md", render_home_page("zh"))
    write_file(DOCS_DIR / "index.en.md", render_home_page("en"))

    for section_slug, section_pages in pages_by_section.items():
        write_file(DOCS_DIR / section_slug / "index.md", render_section_index(section_slug, section_pages, manifest, "zh"))
        write_file(DOCS_DIR / section_slug / "index.en.md", render_section_index(section_slug, section_pages, manifest, "en"))
        for page in section_pages:
            write_file(DOCS_DIR / section_slug / f"{page['slug']}.md", render_page(page, manifest, units_by_page, units_by_source, "zh"))
            write_file(DOCS_DIR / section_slug / f"{page['slug']}.en.md", render_page(page, manifest, units_by_page, units_by_source, "en"))

    write_file(
        DATA_DIR / "source_units.jsonl",
        "\n".join(json.dumps(unit, ensure_ascii=False) for unit in source_units) + "\n",
    )
    write_file(
        DATA_DIR / "coverage_map.json",
        json.dumps(build_coverage_map(source_units), ensure_ascii=False, indent=2),
    )
    write_file(
        DATA_DIR / "extraction_summary.json",
        json.dumps(
            {
                **summary,
                "pages": {
                    page["slug"]: {
                        "section": page["section"],
                        "source_files": page["source_files"],
                    }
                    for page in PAGES
                },
                "sources": manifest,
            },
            ensure_ascii=False,
            indent=2,
        ),
    )


def main() -> None:
    ensure_clean_generation_dirs()
    source_units, manifest, warnings, summary = extract_all_sources()
    build_docs(source_units, manifest, summary)
    print(f"Generated HFE notes site with {len(manifest)} source files and {len(source_units)} text units.")
    if warnings:
        print("Warnings:")
        for warning in warnings:
            print(f"- {warning}")


if __name__ == "__main__":
    main()
